from __future__ import annotations

import asyncio
import json
import tempfile
import unittest
from pathlib import Path

import aiohttp
from aiohttp import web
from openpyxl import load_workbook

from llm_load_tester.benchmarker import LLMBenchmarker, LoadTestConfig, SSEDecoder
from llm_load_tester.metrics import ErrorCategory
from llm_load_tester.modelarts_monitoring import (
    ModelArtsCloudEyeCollector,
    ModelArtsCloudEyeConfig,
)
from llm_load_tester.modalities import ModalityHandler, PayloadResult


class FixedHandler(ModalityHandler):
    async def prepare_payload(self, config: dict) -> PayloadResult:
        return PayloadResult(
            payload={
                "model": config["model"],
                "messages": [{"role": "user", "content": "hello"}],
                "stream": True,
            },
            metadata={"input_tokens": 4},
        )

    def get_presets(self) -> dict:
        return {}


class BenchmarkIntegrationTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self) -> None:
        self.state = {
            "running": 0,
            "prompt_tokens": 50,
            "generation_tokens": 100,
            "successful_requests": 5,
            "ttft_count": 5,
            "ttft_sum": 0.5,
            "ttft_le_point_one": 4,
        }
        app = web.Application()
        app.router.add_post("/stream", self._stream)
        app.router.add_post("/buffered", self._buffered)
        app.router.add_post("/scenario", self._scenario)
        app.router.add_get("/metrics", self._metrics)
        app.router.add_post(
            "/V1.0/{project_id}/batch-query-metric-data", self._cloud_eye_metrics
        )
        self.runner = web.AppRunner(app)
        await self.runner.setup()
        self.site = web.TCPSite(self.runner, "127.0.0.1", 0)
        await self.site.start()
        sockets = self.site._server.sockets
        self.endpoint = f"http://127.0.0.1:{sockets[0].getsockname()[1]}"

    async def asyncTearDown(self) -> None:
        await self.runner.cleanup()

    def _config(self, route: str, **overrides) -> LoadTestConfig:
        values = {
            "endpoint": self.endpoint,
            "api_route": route,
            "api_key": None,
            "model": "mock-model",
            "concurrency": 2,
            "target_rps": 0,
            "total_requests": 2,
            "warmup_requests": 0,
            "modality_handler": FixedHandler(),
            "modality_config": {"modality": "text", "model": "mock-model"},
            "timeout_seconds": 5.0,
        }
        values.update(overrides)
        return LoadTestConfig(**values)

    async def _write_event(self, response: web.StreamResponse, payload: dict) -> None:
        await response.write(f"data: {json.dumps(payload)}\n\n".encode())

    async def _stream(self, request: web.Request) -> web.StreamResponse:
        response = web.StreamResponse(
            headers={
                "Content-Type": "text/event-stream",
                "X-Request-Id": "stream-1",
                "X-Span-Id": "span-1",
            }
        )
        await response.prepare(request)
        await response.write(b": heartbeat\n\n")
        await self._write_event(response, {
            "id": "completion-1",
            "model": "glm-5.2",
            "object": "chat.completion.chunk",
            "service_tier": "default",
            "created": 1000,
            "first_token_return_time": 1000.0,
            "choices": [{"delta": {"role": "assistant"}}],
        })
        await asyncio.sleep(0.01)
        await self._write_event(
            response, {
                "first_token_return_time": 1000.01,
                "choices": [{"delta": {"reasoning_content": "Think"}}],
            }
        )
        await asyncio.sleep(0.01)
        await self._write_event(response, {
            "first_token_return_time": 1000.02,
            "choices": [{"delta": {"content": "Hello"}, "finish_reason": "stop"}],
        })
        await self._write_event(
            response,
            {
                "first_token_return_time": 1000.03,
                "choices": [],
                "usage": {
                    "prompt_tokens": 4,
                    "completion_tokens": 3,
                    "total_tokens": 7,
                    "prompt_tokens_details": {"cached_tokens": 1},
                    "completion_tokens_details": {"reasoning_tokens": 2},
                },
            },
        )
        await response.write(b"data: [DONE]\n\n")
        await response.write_eof()
        return response

    async def _buffered(self, request: web.Request) -> web.Response:
        await asyncio.sleep(0.01)
        return web.json_response(
            {
                "choices": [{"message": {"role": "assistant", "content": "Hello"}}],
                "usage": {"prompt_tokens": 4, "completion_tokens": 2},
            },
            headers={"X-Request-Id": "buffered-1", "X-Span-Id": "buffered-span"},
        )

    async def _scenario(self, request: web.Request) -> web.StreamResponse:
        self.state["running"] += 1
        response = web.StreamResponse(headers={"Content-Type": "text/event-stream"})
        await response.prepare(request)
        await self._write_event(response, {"choices": [{"delta": {"role": "assistant"}}]})
        await asyncio.sleep(0.12)
        await self._write_event(response, {"choices": [{"delta": {"content": "Hello world"}}]})
        await self._write_event(
            response,
            {"choices": [], "usage": {"prompt_tokens": 4, "completion_tokens": 12}},
        )
        self.state["running"] -= 1
        self.state["prompt_tokens"] += 4
        self.state["generation_tokens"] += 12
        self.state["successful_requests"] += 1
        self.state["ttft_count"] += 1
        self.state["ttft_sum"] += 0.05
        self.state["ttft_le_point_one"] += 1
        await response.write(b"data: [DONE]\n\n")
        await response.write_eof()
        return response

    async def _metrics(self, request: web.Request) -> web.Response:
        state = self.state
        body = f"""
# HELP vllm:num_requests_running Running requests.
# TYPE vllm:num_requests_running gauge
vllm:num_requests_running{{model_name="mock-model"}} {state['running']}
# HELP vllm:prompt_tokens_total Prompt tokens.
# TYPE vllm:prompt_tokens_total counter
vllm:prompt_tokens_total{{model_name="mock-model"}} {state['prompt_tokens']}
# HELP vllm:generation_tokens_total Generated tokens.
# TYPE vllm:generation_tokens_total counter
vllm:generation_tokens_total{{model_name="mock-model"}} {state['generation_tokens']}
# HELP vllm:request_success_total Successful requests.
# TYPE vllm:request_success_total counter
vllm:request_success_total{{model_name="mock-model",finished_reason="stop"}} {state['successful_requests']}
# HELP vllm:time_to_first_token_seconds Time to first token.
# TYPE vllm:time_to_first_token_seconds histogram
vllm:time_to_first_token_seconds_bucket{{model_name="mock-model",le="0.1"}} {state['ttft_le_point_one']}
vllm:time_to_first_token_seconds_bucket{{model_name="mock-model",le="+Inf"}} {state['ttft_count']}
vllm:time_to_first_token_seconds_sum{{model_name="mock-model"}} {state['ttft_sum']}
vllm:time_to_first_token_seconds_count{{model_name="mock-model"}} {state['ttft_count']}
"""
        return web.Response(text=body, content_type="text/plain")

    async def _cloud_eye_metrics(self, request: web.Request) -> web.Response:
        self.cloud_eye_request = {
            "project_id": request.match_info["project_id"],
            "token": request.headers.get("X-Auth-Token"),
            "body": await request.json(),
        }
        return web.json_response({
            "metrics": [
                {
                    "namespace": "SYS.MaaS",
                    "metric_name": "ttft",
                    "dimensions": [{"name": "maas_api_id", "value": "api-1"}],
                    "unit": "ms",
                    "datapoints": [{"average": 2500.0, "timestamp": 1000}],
                },
                {
                    "namespace": "SYS.MaaS",
                    "metric_name": "req_count",
                    "dimensions": [{"name": "maas_api_id", "value": "api-1"}],
                    "unit": "count/min",
                    "datapoints": [{"sum": 2.0, "timestamp": 1000}],
                },
            ]
        })

    async def test_streaming_response_reports_ttft_and_true_ttfb(self) -> None:
        benchmarker = LLMBenchmarker(self._config("stream", total_requests=1))
        async with aiohttp.ClientSession() as session:
            metric = await benchmarker._make_request(session, request_id=1)

        self.assertEqual(metric.error, ErrorCategory.NONE)
        self.assertEqual(metric.response_mode, "streaming")
        self.assertEqual(metric.responsiveness_metric_type, "ttft")
        self.assertEqual(metric.available_responsiveness_metrics, ["ttfb", "ttft"])
        self.assertIsNotNone(metric.ttfb)
        self.assertIsNotNone(metric.ttft)
        self.assertLess(metric.ttfb, metric.ttft)
        self.assertEqual(metric.first_output_kind, "reasoning")
        self.assertIsNotNone(metric.time_to_first_reasoning)
        self.assertIsNotNone(metric.time_to_first_text)
        self.assertEqual(metric.reasoning_content, "Think")
        self.assertEqual(metric.response_content, "Hello")
        self.assertEqual(metric.tokens_generated, 3)
        self.assertEqual(metric.token_count_source, "provider_usage")
        self.assertEqual(metric.upstream_request_id, "stream-1")
        self.assertEqual(metric.upstream_span_id, "span-1")
        self.assertEqual(metric.provider_response_id, "completion-1")
        self.assertEqual(metric.provider_model, "glm-5.2")
        self.assertEqual(metric.provider_service_tier, "default")
        self.assertEqual(metric.provider_first_token_return_time, 1000.0)
        self.assertEqual(metric.provider_last_token_return_time, 1000.03)
        self.assertAlmostEqual(metric.provider_stream_return_span, 0.03)
        self.assertEqual(metric.provider_usage["completion_tokens_details"]["reasoning_tokens"], 2)
        self.assertEqual(metric.finish_reason, "stop")
        benchmarker.result.add_request(metric)
        summary = benchmarker.result.get_summary()
        self.assertEqual(summary["total_reasoning_tokens"], 2)
        self.assertEqual(summary["total_cached_input_tokens"], 1)

    async def test_buffered_response_selects_ttfb_and_does_not_invent_ttft(self) -> None:
        benchmarker = LLMBenchmarker(self._config("buffered", total_requests=1))
        async with aiohttp.ClientSession() as session:
            metric = await benchmarker._make_request(session, request_id=1)

        self.assertEqual(metric.error, ErrorCategory.NONE)
        self.assertEqual(metric.response_mode, "buffered")
        self.assertEqual(metric.responsiveness_metric_type, "ttfb")
        self.assertIsNotNone(metric.ttfb)
        self.assertIsNone(metric.ttft)
        self.assertEqual(metric.available_responsiveness_metrics, ["ttfb"])
        self.assertEqual(metric.tokens_generated, 2)
        self.assertEqual(metric.upstream_span_id, "buffered-span")

    async def test_modelarts_cloud_eye_adapter_is_separate_and_normalizes_latency(self) -> None:
        collector = ModelArtsCloudEyeCollector(ModelArtsCloudEyeConfig(
            endpoint=self.endpoint,
            project_id="project-1",
            iam_token="iam-token",
            dimensions=(("maas_api_id", "api-1"),),
            metric_names=("ttft", "req_count"),
        ))
        async with aiohttp.ClientSession() as session:
            result = await collector.collect(session, 100_000, 200_000)

        self.assertTrue(result["available"])
        self.assertEqual(result["provider"], "modelarts_cloud_eye")
        self.assertEqual(result["metrics"][0]["latest"]["average"], 2.5)
        self.assertEqual(result["metrics"][0]["unit"], "seconds")
        self.assertEqual(result["metrics"][1]["latest"]["sum"], 2.0)
        self.assertEqual(self.cloud_eye_request["project_id"], "project-1")
        self.assertEqual(self.cloud_eye_request["token"], "iam-token")
        self.assertEqual(self.cloud_eye_request["body"]["metrics"][0]["namespace"], "SYS.MaaS")

    async def test_scenario_metrics_are_collected_without_a_gpu(self) -> None:
        benchmarker = LLMBenchmarker(
            self._config(
                "scenario",
                metrics_url=f"{self.endpoint}/metrics",
                metrics_scrape_interval_seconds=0.05,
                modelarts_cloud_eye=ModelArtsCloudEyeConfig(
                    endpoint=self.endpoint,
                    project_id="project-1",
                    iam_token="iam-token",
                    dimensions=(("maas_api_id", "api-1"),),
                    metric_names=("ttft", "req_count"),
                ),
            )
        )
        result = await benchmarker.run()

        self.assertEqual(result.successful_requests, 2)
        self.assertIsNotNone(result.server_metrics)
        server_metrics = result.server_metrics
        self.assertTrue(server_metrics["available"])
        self.assertGreaterEqual(server_metrics["scrape_count"], 3)
        self.assertEqual(server_metrics["elapsed_reference"], "measured_scenario_start")
        summary = server_metrics["summary"]
        self.assertEqual(summary["counters"]["generation_tokens"]["delta"], 24.0)
        self.assertEqual(summary["counters"]["prompt_tokens"]["delta"], 8.0)
        self.assertEqual(summary["gauges"]["running_requests"]["max"], 2.0)
        self.assertEqual(summary["histograms"]["ttft"]["count_delta"], 2.0)
        self.assertAlmostEqual(summary["histograms"]["ttft"]["mean"], 0.05)

        client_summary = result.get_summary()
        self.assertEqual(client_summary["responsiveness_metric_type"], "ttft")
        self.assertEqual(client_summary["provider_responsiveness_metric_type"], "ttft")
        self.assertIsNotNone(result.provider_monitoring)
        self.assertTrue(result.provider_monitoring["available"])
        self.assertEqual(
            result.provider_monitoring["metrics"][0]["latest"]["average"], 2.5
        )

        with tempfile.TemporaryDirectory() as directory:
            csv_path = Path(directory) / "result.csv"
            json_path = Path(directory) / "result.json"
            xlsx_path = Path(directory) / "result.xlsx"
            result.export_csv(csv_path)
            result.export_json(json_path)
            result.export_xlsx(xlsx_path)
            self.assertTrue(csv_path.exists())
            self.assertTrue((Path(directory) / "result.vllm_metrics.csv").exists())
            self.assertTrue((Path(directory) / "result.vllm_metrics.json").exists())
            self.assertTrue((Path(directory) / "result.modelarts_metrics.csv").exists())
            self.assertTrue((Path(directory) / "result.modelarts_metrics.json").exists())
            exported = json.loads(json_path.read_text())
            self.assertIn("server_metrics", exported)
            self.assertIn("provider_monitoring", exported)
            self.assertEqual(
                exported["summary_units"]["client_observed_ttft_min"],
                "seconds",
            )
            self.assertEqual(
                exported["summary_units"]["overall_tokens_per_second"],
                "tokens/second",
            )
            self.assertEqual(exported["summary_units"]["tpot_mean"], "seconds/token")

            workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Benchmark Summary",
                    "Requests",
                    "vLLM Summary",
                    "vLLM Timeline",
                    "Request Context",
                    "ModelArts Metrics",
                ],
            )
            self.assertEqual(workbook["Requests"].max_row, 3)
            benchmark_rows = {
                row[0]: row[1:]
                for row in workbook["Benchmark Summary"].iter_rows(
                    min_row=5,
                    values_only=True,
                )
            }
            self.assertEqual(
                benchmark_rows["client_observed_ttft_min"][1],
                "seconds",
            )
            self.assertEqual(
                benchmark_rows["error_rate_percent"][1],
                "percent",
            )
            self.assertEqual(
                benchmark_rows["overall_tokens_per_second"][1],
                "tokens/second",
            )
            self.assertEqual(benchmark_rows["tpot_mean"][1], "seconds/token")
            vllm_rows = list(workbook["vLLM Summary"].iter_rows(values_only=True))
            self.assertIn(
                ("counter", "generation_tokens", "delta", 24, "tokens"),
                [row[:5] for row in vllm_rows[1:]],
            )
            context_rows = list(workbook["Request Context"].iter_rows(values_only=True))
            self.assertEqual(len(context_rows), 3)
            self.assertTrue(
                all(row[14] in {"within_request_window", "nearest_scrape"} for row in context_rows[1:])
            )
            self.assertTrue(
                all("not metrics attributed" in row[22] for row in context_rows[1:])
            )
            workbook.close()

    async def test_unavailable_metrics_endpoint_does_not_fail_load_by_default(self) -> None:
        benchmarker = LLMBenchmarker(
            self._config(
                "stream",
                total_requests=1,
                concurrency=1,
                metrics_url=f"{self.endpoint}/missing-metrics",
                metrics_scrape_interval_seconds=0.05,
            )
        )

        result = await benchmarker.run()

        self.assertEqual(result.successful_requests, 1)
        self.assertFalse(result.server_metrics["available"])
        self.assertGreaterEqual(len(result.server_metrics["errors"]), 2)
        self.assertEqual(result.server_metrics["summary"], {})

    def test_sse_decoder_handles_split_utf8_and_multiline_events(self) -> None:
        decoder = SSEDecoder()
        encoded = 'data: {"text":"ı"}\ndata: second\n\n'.encode("utf-8")
        events = decoder.feed(encoded[:-2])
        events.extend(decoder.feed(encoded[-2:]))
        events.extend(decoder.finalize())
        self.assertEqual(events, ['{"text":"ı"}\nsecond'])


if __name__ == "__main__":
    unittest.main()
