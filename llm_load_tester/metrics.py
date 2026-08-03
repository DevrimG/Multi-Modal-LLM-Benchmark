from __future__ import annotations

"""
Metrics collection, calculation, and export for LLM load testing.
"""

import csv
import json
import statistics
from dataclasses import dataclass, field, asdict
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Optional

import numpy as np
from rich.console import Console
from rich.table import Table
from tabulate import tabulate


class ErrorCategory(Enum):
    """Categories of errors that can occur during testing."""
    NONE = "none"
    RATE_LIMIT = "rate_limit"  # 429
    TIMEOUT = "timeout"
    CONNECTION_ERROR = "connection_error"
    SERVER_ERROR = "server_error"  # 5xx
    CLIENT_ERROR = "client_error"  # 4xx
    UNKNOWN = "unknown"
    
    @classmethod
    def from_status_code(cls, status_code: int | None) -> "ErrorCategory":
        """Categorize error based on HTTP status code."""
        if status_code is None:
            return cls.CONNECTION_ERROR
        if status_code == 429:
            return cls.RATE_LIMIT
        if 500 <= status_code < 600:
            return cls.SERVER_ERROR
        if 400 <= status_code < 500:
            return cls.CLIENT_ERROR
        return cls.UNKNOWN


@dataclass
class RequestMetrics:
    """Metrics for a single request."""
    request_id: int
    start_time: float
    headers_received_time: float | None = None
    first_byte_time: float | None = None
    first_event_time: float | None = None
    first_output_time: float | None = None
    first_token_time: float | None = None
    first_reasoning_time: float | None = None
    first_text_time: float | None = None
    first_tool_call_time: float | None = None
    first_audio_time: float | None = None
    end_time: float | None = None
    tokens_generated: int = 0
    input_tokens: int = 0
    error: ErrorCategory = ErrorCategory.NONE
    error_message: str = ""
    status_code: int | None = None
    response_content: str = ""
    reasoning_content: str = ""
    upstream_request_id: str | None = None
    upstream_span_id: str | None = None
    provider_response_id: str | None = None
    provider_model: str | None = None
    provider_object: str | None = None
    provider_service_tier: str | None = None
    provider_created: float | None = None
    provider_first_token_return_time: float | None = None
    provider_last_token_return_time: float | None = None
    provider_usage: dict[str, Any] = field(default_factory=dict)
    finish_reason: str | None = None
    first_output_kind: str | None = None
    response_mode: str | None = None
    response_content_type: str = ""
    token_count_source: str = "character_estimate"

    @property
    def ttfb(self) -> float | None:
        """Client-observed time to first non-empty response-body byte."""
        if self.first_byte_time is not None:
            return self.first_byte_time - self.start_time
        return None

    @property
    def time_to_first_event(self) -> float | None:
        """Time to the first complete SSE data event."""
        if self.first_event_time is not None:
            return self.first_event_time - self.start_time
        return None

    @property
    def time_to_first_output(self) -> float | None:
        """Time to the first identifiable output-bearing response object."""
        if self.first_output_time is not None:
            return self.first_output_time - self.start_time
        return None

    @property
    def time_to_first_reasoning(self) -> float | None:
        if self.first_reasoning_time is not None:
            return self.first_reasoning_time - self.start_time
        return None

    @property
    def time_to_first_text(self) -> float | None:
        if self.first_text_time is not None:
            return self.first_text_time - self.start_time
        return None

    @property
    def time_to_first_tool_call(self) -> float | None:
        if self.first_tool_call_time is not None:
            return self.first_tool_call_time - self.start_time
        return None

    @property
    def time_to_first_audio(self) -> float | None:
        if self.first_audio_time is not None:
            return self.first_audio_time - self.start_time
        return None
    
    @property
    def ttft(self) -> float | None:
        """Time to first identifiable token-bearing streamed output."""
        if self.first_token_time is not None:
            return self.first_token_time - self.start_time
        return None
    
    @property
    def total_latency(self) -> float | None:
        """Client-observed total request latency in seconds."""
        if self.end_time is not None:
            return self.end_time - self.start_time
        return None

    @property
    def responsiveness_metric_type(self) -> str:
        """Primary comparable responsiveness metric selected from response evidence."""
        if self.response_mode == "streaming" and self.ttft is not None:
            return "ttft"
        if self.ttfb is not None:
            return "ttfb"
        if self.ttft is not None:
            return "ttft"
        return "unavailable"

    @property
    def available_responsiveness_metrics(self) -> list[str]:
        available: list[str] = []
        if self.ttfb is not None:
            available.append("ttfb")
        if self.ttft is not None:
            available.append("ttft")
        return available

    @property
    def provider_stream_return_span(self) -> float | None:
        """Provider timestamp span across returned chunks; not a TTFT measurement."""
        if (
            self.provider_first_token_return_time is not None
            and self.provider_last_token_return_time is not None
        ):
            return self.provider_last_token_return_time - self.provider_first_token_return_time
        return None

    @property
    def measurement_basis(self) -> str:
        if self.response_mode == "streaming" and self.ttft is not None:
            return "streamed_token_bearing_delta"
        if self.response_mode == "buffered" and self.ttfb is not None:
            return "buffered_response_ttfb_only"
        if self.ttfb is not None:
            return "response_body_bytes"
        return "unavailable"
    
    @property
    def tpot(self) -> float | None:
        """Time Per Output Token in seconds."""
        if self.tokens_generated > 0 and self.total_latency is not None:
            # Exclude TTFT from TPOT calculation
            if self.ttft is not None:
                return (self.total_latency - self.ttft) / self.tokens_generated
            return self.total_latency / self.tokens_generated
        return None
    
    @property
    def tokens_per_second(self) -> float | None:
        """Tokens per second for this request."""
        if self.tokens_generated > 0 and self.total_latency is not None:
            return self.tokens_generated / self.total_latency
        return None
    
    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary for export."""
        return {
            "request_id": self.request_id,
            "upstream_request_id": self.upstream_request_id,
            "upstream_span_id": self.upstream_span_id,
            "provider_response_id": self.provider_response_id,
            "provider_model": self.provider_model,
            "provider_object": self.provider_object,
            "provider_service_tier": self.provider_service_tier,
            "provider_created": self.provider_created,
            "provider_first_token_return_time": self.provider_first_token_return_time,
            "provider_last_token_return_time": self.provider_last_token_return_time,
            "provider_stream_return_span_seconds": self.provider_stream_return_span,
            "provider_timestamp_note": (
                "Provider absolute chunk-return timestamps; not client or server TTFT."
                if self.provider_first_token_return_time is not None
                else None
            ),
            "provider_usage": self.provider_usage,
            "finish_reason": self.finish_reason,
            "start_time": self.start_time,
            "headers_received_time": self.headers_received_time,
            "first_byte_time": self.first_byte_time,
            "first_event_time": self.first_event_time,
            "first_output_time": self.first_output_time,
            "first_token_time": self.first_token_time,
            "first_reasoning_time": self.first_reasoning_time,
            "first_text_time": self.first_text_time,
            "first_tool_call_time": self.first_tool_call_time,
            "first_audio_time": self.first_audio_time,
            "end_time": self.end_time,
            "client_observed_ttfb_seconds": self.ttfb,
            "client_observed_time_to_first_event_seconds": self.time_to_first_event,
            "client_observed_time_to_first_output_seconds": self.time_to_first_output,
            "client_observed_ttft_seconds": self.ttft,
            "ttft_seconds": self.ttft,
            "client_observed_time_to_first_reasoning_seconds": self.time_to_first_reasoning,
            "client_observed_time_to_first_text_seconds": self.time_to_first_text,
            "client_observed_time_to_first_tool_call_seconds": self.time_to_first_tool_call,
            "client_observed_time_to_first_audio_seconds": self.time_to_first_audio,
            "client_observed_latency_seconds": self.total_latency,
            "total_latency_seconds": self.total_latency,
            "tpot_seconds_per_token": self.tpot,
            "tpot_seconds": self.tpot,
            "tokens_generated": self.tokens_generated,
            "input_tokens": self.input_tokens,
            "tokens_per_second": self.tokens_per_second,
            "token_count_source": self.token_count_source,
            "responsiveness_metric_type": self.responsiveness_metric_type,
            "available_responsiveness_metrics": self.available_responsiveness_metrics,
            "measurement_basis": self.measurement_basis,
            "first_output_kind": self.first_output_kind,
            "response_mode": self.response_mode,
            "response_content_type": self.response_content_type,
            "error": self.error.value,
            "error_message": self.error_message,
            "status_code": self.status_code,
            "response_content": self.response_content,
            "reasoning_content": self.reasoning_content,
        }


@dataclass
class BenchmarkResult:
    """Aggregated results from a benchmark run."""
    # Configuration
    modality: str
    model: str
    endpoint: str
    concurrency: int
    target_rps: float
    total_requests: int
    warmup_requests: int
    
    # Timing
    start_time: datetime
    end_time: datetime | None = None
    
    # Text modality configuration
    input_tokens: int | None = None
    output_tokens: int | None = None
    output_token_parameter: str | None = None
    thinking_mode: str | None = None
    
    # Image modality configuration
    image_directory: str | None = None
    
    # Voice modality configuration
    audio_directory: str | None = None
    audio_file: str | None = None
    
    # Raw metrics
    request_metrics: list[RequestMetrics] = field(default_factory=list)
    server_metrics: dict[str, Any] | None = None
    provider_monitoring: dict[str, Any] | None = None
    scenario_start_monotonic: float | None = field(default=None, repr=False)
    
    # Error tracking
    errors: dict[ErrorCategory, int] = field(default_factory=lambda: {
        cat: 0 for cat in ErrorCategory
    })
    
    def add_request(self, metrics: RequestMetrics) -> None:
        """Add a request's metrics to the results."""
        self.request_metrics.append(metrics)
        if metrics.error != ErrorCategory.NONE:
            self.errors[metrics.error] += 1
    
    @property
    def successful_requests(self) -> int:
        """Count of successful requests."""
        return sum(1 for m in self.request_metrics if m.error == ErrorCategory.NONE)
    
    @property
    def failed_requests(self) -> int:
        """Count of failed requests."""
        return sum(1 for m in self.request_metrics if m.error != ErrorCategory.NONE)
    
    @property
    def error_rate(self) -> float:
        """Error rate as a percentage."""
        if not self.request_metrics:
            return 0.0
        return (self.failed_requests / len(self.request_metrics)) * 100
    
    def _calculate_percentile(self, values: list[float], percentile: float) -> float | None:
        """Calculate percentile value from a list of numbers."""
        if not values:
            return None
        return np.percentile(values, percentile)
    
    def _get_valid_values(self, extractor) -> list[float]:
        """Extract valid (non-None) values from request metrics."""
        return [v for m in self.request_metrics if (v := extractor(m)) is not None]
    
    def get_summary(self) -> dict[str, Any]:
        """Generate a summary of benchmark results."""
        # Collect all valid values
        successful_metrics = [
            metric for metric in self.request_metrics if metric.error == ErrorCategory.NONE
        ]
        client_ttfbs = [
            value for metric in successful_metrics if (value := metric.ttfb) is not None
        ]
        client_ttfts = [
            value for metric in successful_metrics if (value := metric.ttft) is not None
        ]
        client_first_text = [
            value
            for metric in successful_metrics
            if (value := metric.time_to_first_text) is not None
        ]
        latencies = self._get_valid_values(lambda m: m.total_latency)
        tpots = self._get_valid_values(lambda m: m.tpot)
        tps_per_req = self._get_valid_values(lambda m: m.tokens_per_second)
        
        # Total tokens
        total_output_tokens = sum(m.tokens_generated for m in self.request_metrics)
        total_input_tokens = sum(m.input_tokens for m in self.request_metrics)
        total_reasoning_tokens = sum(
            int(
                ((m.provider_usage.get("completion_tokens_details") or {}).get("reasoning_tokens"))
                or 0
            )
            for m in self.request_metrics
        )
        total_cached_input_tokens = sum(
            int(
                ((m.provider_usage.get("prompt_tokens_details") or {}).get("cached_tokens"))
                or 0
            )
            for m in self.request_metrics
        )
        
        # Overall duration
        duration_seconds = None
        if self.end_time and self.start_time:
            duration_seconds = (self.end_time - self.start_time).total_seconds()
        
        # Overall throughput
        overall_tps = None
        if duration_seconds and duration_seconds > 0:
            overall_tps = total_output_tokens / duration_seconds

        successful_count = self.successful_requests
        ttfb_coverage = (len(client_ttfbs) / successful_count * 100) if successful_count else 0.0
        ttft_coverage = (len(client_ttfts) / successful_count * 100) if successful_count else 0.0
        primary_metric_types = {
            metric.responsiveness_metric_type for metric in successful_metrics
        }
        primary_metric_types.discard("unavailable")
        if len(primary_metric_types) == 1:
            responsiveness_type = next(iter(primary_metric_types))
        elif len(primary_metric_types) > 1:
            responsiveness_type = "mixed"
        else:
            responsiveness_type = "unavailable"

        provider_histograms = (
            ((self.server_metrics or {}).get("summary") or {}).get("histograms") or {}
        )
        provider_ttft = provider_histograms.get("ttft") or {}
        provider_ttfb = provider_histograms.get("ttfb") or {}
        if provider_ttft and provider_ttfb:
            provider_responsiveness_type = "both"
        elif provider_ttft:
            provider_responsiveness_type = "ttft"
        elif provider_ttfb:
            provider_responsiveness_type = "ttfb"
        else:
            provider_responsiveness_type = "unavailable"
        
        summary = {
            # Configuration
            "modality": self.modality,
            "model": self.model,
            "endpoint": self.endpoint,
            "concurrency": self.concurrency,
            "target_rps": self.target_rps,
            "total_requests": self.total_requests,
            "warmup_requests": self.warmup_requests,
            "actual_requests": len(self.request_metrics),
            
            # Modality-specific configuration
            "input_tokens": self.input_tokens,
            "output_tokens": self.output_tokens,
            "output_token_parameter": self.output_token_parameter,
            "thinking_mode": self.thinking_mode,
            "image_directory": self.image_directory,
            "audio_directory": self.audio_directory,
            "audio_file": self.audio_file,
            
            # Timing
            "duration_seconds": duration_seconds,
            "start_time": self.start_time.isoformat(),
            "end_time": self.end_time.isoformat() if self.end_time else None,
            
            # Success metrics
            "successful_requests": self.successful_requests,
            "failed_requests": self.failed_requests,
            "error_rate_percent": round(self.error_rate, 2),
            
            # Token counts
            "total_input_tokens": total_input_tokens,
            "total_output_tokens": total_output_tokens,
            "total_tokens": total_input_tokens + total_output_tokens,
            "total_reasoning_tokens": total_reasoning_tokens,
            "total_cached_input_tokens": total_cached_input_tokens,

            # Responsiveness metric availability
            "responsiveness_metric_type": responsiveness_type,
            "provider_responsiveness_metric_type": provider_responsiveness_type,
            "ttfb_coverage_percent": round(ttfb_coverage, 2),
            "ttft_coverage_percent": round(ttft_coverage, 2),

            # Client-observed TTFB
            "client_observed_ttfb_mean": round(statistics.mean(client_ttfbs), 3) if client_ttfbs else None,
            "client_observed_ttfb_p50": round(statistics.median(client_ttfbs), 3) if client_ttfbs else None,
            "client_observed_ttfb_p95": round(self._calculate_percentile(client_ttfbs, 95), 3) if client_ttfbs else None,
            "client_observed_ttfb_p99": round(self._calculate_percentile(client_ttfbs, 99), 3) if client_ttfbs else None,
            "client_observed_ttfb_min": round(min(client_ttfbs), 3) if client_ttfbs else None,
            "client_observed_ttfb_max": round(max(client_ttfbs), 3) if client_ttfbs else None,
            
            # Client-observed TTFT. This includes client->gateway transit,
            # provider edge/gateway handling, server-side queueing, and any
            # buffering before the first streamed text chunk reaches the client.
            "client_observed_ttft_mean": round(statistics.mean(client_ttfts), 3) if client_ttfts else None,
            "client_observed_ttft_p50": round(statistics.median(client_ttfts), 3) if client_ttfts else None,
            "client_observed_ttft_p95": round(self._calculate_percentile(client_ttfts, 95), 3) if client_ttfts else None,
            "client_observed_ttft_p99": round(self._calculate_percentile(client_ttfts, 99), 3) if client_ttfts else None,
            "client_observed_ttft_min": round(min(client_ttfts), 3) if client_ttfts else None,
            "client_observed_ttft_max": round(max(client_ttfts), 3) if client_ttfts else None,
            "client_observed_time_to_first_text_mean": round(statistics.mean(client_first_text), 3) if client_first_text else None,
            "provider_internal_ttft_mean": provider_ttft.get("mean"),
            "provider_internal_ttft_p50": provider_ttft.get("p50"),
            "provider_internal_ttft_p95": provider_ttft.get("p95"),
            "provider_internal_ttft_p99": provider_ttft.get("p99"),
            "provider_internal_ttft_source": provider_ttft.get("source_metrics"),
            "provider_internal_ttft_note": (
                "Scenario-level histogram delta from the server metrics endpoint."
                if provider_ttft
                else "Not available from the configured server metrics endpoint."
            ),
            "provider_internal_ttfb_mean": provider_ttfb.get("mean"),
            "provider_internal_ttfb_p50": provider_ttfb.get("p50"),
            "provider_internal_ttfb_p95": provider_ttfb.get("p95"),
            "provider_internal_ttfb_p99": provider_ttfb.get("p99"),
            "provider_internal_ttfb_source": provider_ttfb.get("source_metrics"),
            # Backward-compatible aliases. Prefer client_observed_ttft_* in new reports.
            "ttft_mean": round(statistics.mean(client_ttfts), 3) if client_ttfts else None,
            "ttft_p50": round(statistics.median(client_ttfts), 3) if client_ttfts else None,
            "ttft_p95": round(self._calculate_percentile(client_ttfts, 95), 3) if client_ttfts else None,
            "ttft_p99": round(self._calculate_percentile(client_ttfts, 99), 3) if client_ttfts else None,
            "ttft_min": round(min(client_ttfts), 3) if client_ttfts else None,
            "ttft_max": round(max(client_ttfts), 3) if client_ttfts else None,
            
            # End-to-End Latency
            "latency_mean": round(statistics.mean(latencies), 3) if latencies else None,
            "latency_p50": round(statistics.median(latencies), 3) if latencies else None,
            "latency_p95": round(self._calculate_percentile(latencies, 95), 3) if latencies else None,
            "latency_p99": round(self._calculate_percentile(latencies, 99), 3) if latencies else None,
            "latency_min": round(min(latencies), 3) if latencies else None,
            "latency_max": round(max(latencies), 3) if latencies else None,
            
            # TPOT (Time Per Output Token)
            "tpot_mean": round(statistics.mean(tpots), 4) if tpots else None,
            "tpot_p50": round(statistics.median(tpots), 4) if tpots else None,
            "tpot_p95": round(self._calculate_percentile(tpots, 95), 4) if tpots else None,
            "tpot_p99": round(self._calculate_percentile(tpots, 99), 4) if tpots else None,
            
            # Throughput
            "overall_tokens_per_second": round(overall_tps, 2) if overall_tps else None,
            "per_request_tokens_per_second_mean": round(statistics.mean(tps_per_req), 2) if tps_per_req else None,
            "per_request_tokens_per_second_p50": round(statistics.median(tps_per_req), 2) if tps_per_req else None,
            
            # Errors
            "errors_by_category": {
                cat.value: count for cat, count in self.errors.items() if count > 0
            }
        }
        
        return summary

    def get_summary_units(self) -> dict[str, str]:
        """Return explicit units for numeric summary fields without renaming keys."""
        units: dict[str, str] = {}
        for metric_name in self.get_summary():
            if metric_name in {"start_time", "end_time"}:
                units[metric_name] = "ISO 8601 timestamp"
            elif metric_name == "duration_seconds":
                units[metric_name] = "seconds"
            elif metric_name == "concurrency":
                units[metric_name] = "concurrent requests"
            elif metric_name == "target_rps":
                units[metric_name] = "requests/second"
            elif metric_name in {
                "total_requests",
                "warmup_requests",
                "actual_requests",
                "successful_requests",
                "failed_requests",
                "errors_by_category",
            }:
                units[metric_name] = "requests"
            elif metric_name.endswith("_percent"):
                units[metric_name] = "percent"
            elif "tokens_per_second" in metric_name:
                units[metric_name] = "tokens/second"
            elif metric_name in {
                "input_tokens",
                "output_tokens",
                "total_input_tokens",
                "total_output_tokens",
                "total_tokens",
                "total_reasoning_tokens",
                "total_cached_input_tokens",
            }:
                units[metric_name] = "tokens"
            elif metric_name.startswith("tpot_"):
                units[metric_name] = "seconds/token"
            elif (
                metric_name.startswith(
                    (
                        "client_observed_ttfb_",
                        "client_observed_ttft_",
                        "client_observed_time_to_first_text_",
                        "provider_internal_ttft_",
                        "provider_internal_ttfb_",
                        "ttft_",
                        "latency_",
                    )
                )
                and not metric_name.endswith(("_source", "_note"))
            ):
                units[metric_name] = "seconds"
        return units
    
    def print_rich_table(self) -> None:
        """Print results using Rich for formatted terminal output."""
        console = Console()
        summary = self.get_summary()
        
        # Main results table
        table = Table(title="Benchmark Results", show_header=True, header_style="bold magenta")
        table.add_column("Metric", style="cyan")
        table.add_column("Value", style="green")
        
        # Configuration section
        table.add_row("[bold]Configuration[/bold]", "")
        table.add_row("  Modality", summary["modality"])
        table.add_row("  Model", summary["model"])
        table.add_row("  Endpoint", summary["endpoint"])
        table.add_row("  Concurrency", str(summary["concurrency"]))
        table.add_row("  Target RPS", str(summary["target_rps"]))
        table.add_row("  Total Requests", str(summary["total_requests"]))
        table.add_row("  Warmup Requests", str(summary["warmup_requests"]))
        
        # Modality-specific configuration
        if summary["modality"] == "text":
            if summary["input_tokens"]:
                table.add_row("  Input Tokens", str(summary["input_tokens"]))
            if summary["output_tokens"]:
                table.add_row("  Output Tokens", str(summary["output_tokens"]))
            if summary["output_token_parameter"]:
                table.add_row("  Output Limit Field", summary["output_token_parameter"])
            if summary["thinking_mode"]:
                table.add_row("  Thinking Mode", summary["thinking_mode"])
        elif summary["modality"] == "image" and summary["image_directory"]:
            table.add_row("  Image Directory", summary["image_directory"])
        elif summary["modality"] == "voice":
            if summary["audio_directory"]:
                table.add_row("  Sound Directory", summary["audio_directory"])
            elif summary["audio_file"]:
                table.add_row("  Audio File", summary["audio_file"])
        
        table.add_row("", "")
        
        # Summary section
        table.add_row("[bold]Summary[/bold]", "")
        table.add_row("  Duration", f"{summary['duration_seconds']:.2f}s" if summary['duration_seconds'] else "N/A")
        table.add_row("  Successful Requests", str(summary["successful_requests"]))
        table.add_row("  Failed Requests", str(summary["failed_requests"]))
        table.add_row("  Error Rate", f"{summary['error_rate_percent']:.2f}%")
        table.add_row("  Total Output Tokens", str(summary["total_output_tokens"]))
        if summary["total_reasoning_tokens"]:
            table.add_row("  Reasoning Tokens", str(summary["total_reasoning_tokens"]))
        if summary["total_cached_input_tokens"]:
            table.add_row("  Cached Input Tokens", str(summary["total_cached_input_tokens"]))
        table.add_row("  Overall Throughput", f"{summary['overall_tokens_per_second']:.2f} tok/s" if summary["overall_tokens_per_second"] else "N/A")
        table.add_row("", "")
        
        # Client-observed responsiveness sections
        table.add_row("[bold]Responsiveness Classification[/bold]", summary["responsiveness_metric_type"])
        table.add_row("  TTFB Coverage", f"{summary['ttfb_coverage_percent']:.2f}%")
        table.add_row("  TTFT Coverage", f"{summary['ttft_coverage_percent']:.2f}%")
        table.add_row("", "")

        table.add_row("[bold]Client-Observed Time To First Byte (TTFB)[/bold]", "")
        table.add_row("  Mean", f"{summary['client_observed_ttfb_mean']:.3f}s" if summary["client_observed_ttfb_mean"] is not None else "N/A")
        table.add_row("  p50", f"{summary['client_observed_ttfb_p50']:.3f}s" if summary["client_observed_ttfb_p50"] is not None else "N/A")
        table.add_row("  p95", f"{summary['client_observed_ttfb_p95']:.3f}s" if summary["client_observed_ttfb_p95"] is not None else "N/A")
        table.add_row("  p99", f"{summary['client_observed_ttfb_p99']:.3f}s" if summary["client_observed_ttfb_p99"] is not None else "N/A")
        table.add_row("  Scope", "Client request start -> first non-empty response-body byte")
        table.add_row("", "")

        table.add_row("[bold]Client-Observed Time To First Token (TTFT)[/bold]", "")
        table.add_row("  Mean", f"{summary['client_observed_ttft_mean']:.3f}s" if summary["client_observed_ttft_mean"] else "N/A")
        table.add_row("  p50", f"{summary['client_observed_ttft_p50']:.3f}s" if summary["client_observed_ttft_p50"] else "N/A")
        table.add_row("  p95", f"{summary['client_observed_ttft_p95']:.3f}s" if summary["client_observed_ttft_p95"] else "N/A")
        table.add_row("  p99", f"{summary['client_observed_ttft_p99']:.3f}s" if summary["client_observed_ttft_p99"] else "N/A")
        table.add_row("  Range", f"{summary['client_observed_ttft_min']:.3f}s - {summary['client_observed_ttft_max']:.3f}s" if summary["client_observed_ttft_min"] else "N/A")
        table.add_row("  Scope", "Client request start -> first token-bearing streamed output")
        table.add_row("  First Visible Text Mean", f"{summary['client_observed_time_to_first_text_mean']:.3f}s" if summary["client_observed_time_to_first_text_mean"] is not None else "N/A")
        provider_ttft = summary["provider_internal_ttft_mean"]
        table.add_row(
            "[bold]Provider-Internal TTFT[/bold]",
            f"{provider_ttft:.3f}s mean" if provider_ttft is not None else "Not available",
        )
        provider_ttfb = summary["provider_internal_ttfb_mean"]
        table.add_row(
            "[bold]Provider-Internal TTFB[/bold]",
            f"{provider_ttfb:.3f}s mean" if provider_ttfb is not None else "Not available",
        )
        table.add_row("", "")

        if self.server_metrics:
            table.add_row("[bold]Server Metrics Collection[/bold]", "")
            available = self.server_metrics.get("available", False)
            has_data = self.server_metrics.get("has_data", False)
            status = "Available" if available and has_data else "Unavailable"
            table.add_row("  Status", status)
            table.add_row("  Scrapes", str(self.server_metrics.get("scrape_count", 0)))
            errors = self.server_metrics.get("errors", [])
            if errors:
                table.add_row("  Collection Error", str(errors[0])[:180])
            table.add_row("", "")

        if self.provider_monitoring:
            table.add_row("[bold]Provider Monitoring[/bold]", "")
            table.add_row(
                "  Provider", str(self.provider_monitoring.get("provider", "unknown"))
            )
            table.add_row(
                "  Status",
                "Available" if self.provider_monitoring.get("available") else "Unavailable",
            )
            table.add_row(
                "  Granularity",
                "Cloud Eye aggregated (1 minute); separate from request timings",
            )
            errors = self.provider_monitoring.get("errors", [])
            if errors:
                table.add_row("  Collection Error", str(errors[0])[:180])
            table.add_row("", "")
        
        # Latency section
        table.add_row("[bold]End-to-End Latency[/bold]", "")
        table.add_row("  Mean", f"{summary['latency_mean']:.3f}s" if summary["latency_mean"] else "N/A")
        table.add_row("  p50", f"{summary['latency_p50']:.3f}s" if summary["latency_p50"] else "N/A")
        table.add_row("  p95", f"{summary['latency_p95']:.3f}s" if summary["latency_p95"] else "N/A")
        table.add_row("  p99", f"{summary['latency_p99']:.3f}s" if summary["latency_p99"] else "N/A")
        table.add_row("  Range", f"{summary['latency_min']:.3f}s - {summary['latency_max']:.3f}s" if summary["latency_min"] else "N/A")
        table.add_row("", "")
        
        # TPOT section
        table.add_row("[bold]Time Per Output Token (TPOT)[/bold]", "")
        table.add_row("  Mean", f"{summary['tpot_mean']:.4f}s" if summary["tpot_mean"] else "N/A")
        table.add_row("  p50", f"{summary['tpot_p50']:.4f}s" if summary["tpot_p50"] else "N/A")
        table.add_row("  p95", f"{summary['tpot_p95']:.4f}s" if summary["tpot_p95"] else "N/A")
        table.add_row("  p99", f"{summary['tpot_p99']:.4f}s" if summary["tpot_p99"] else "N/A")
        table.add_row("", "")
        
        # Throughput section
        table.add_row("[bold]Throughput[/bold]", "")
        table.add_row("  Overall", f"{summary['overall_tokens_per_second']:.2f} tok/s" if summary["overall_tokens_per_second"] else "N/A")
        table.add_row("  Per Request Mean", f"{summary['per_request_tokens_per_second_mean']:.2f} tok/s" if summary["per_request_tokens_per_second_mean"] else "N/A")
        table.add_row("  Per Request p50", f"{summary['per_request_tokens_per_second_p50']:.2f} tok/s" if summary["per_request_tokens_per_second_p50"] else "N/A")
        
        console.print()
        console.print(table)
        
        # Error breakdown if any
        if summary["errors_by_category"]:
            error_table = Table(title="Error Breakdown", show_header=True, header_style="bold red")
            error_table.add_column("Error Category", style="red")
            error_table.add_column("Count", style="yellow")
            for cat, count in summary["errors_by_category"].items():
                error_table.add_row(cat, str(count))
            console.print()
            console.print(error_table)
        
        console.print()
    
    def export_json(self, filepath: Path) -> None:
        """Export results to JSON file."""
        filepath.parent.mkdir(parents=True, exist_ok=True)
        export_data = {
            "summary": self.get_summary(),
            "summary_units": self.get_summary_units(),
            "raw_metrics": [m.to_dict() for m in self.request_metrics],
            "server_metrics": self.server_metrics,
            "provider_monitoring": self.provider_monitoring,
        }
        
        with open(filepath, "w") as f:
            json.dump(export_data, f, indent=2)

    def export_xlsx(self, filepath: Path) -> None:
        """Export one readable workbook with requests and monitoring on separate sheets."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "Excel export requires openpyxl. Install project dependencies with: "
                "python3.12 -m pip install -r requirements.txt"
            ) from exc

        filepath.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        navy = "1F4E78"
        pale_blue = "D9EAF7"
        pale_gray = "F3F6F8"
        white = "FFFFFF"
        dark_text = "1F2937"
        header_border = Border(bottom=Side(style="thin", color="9CA3AF"))

        def excel_value(value: Any) -> Any:
            if isinstance(value, (dict, list, tuple, set)):
                return json.dumps(value, sort_keys=True, ensure_ascii=False)
            if isinstance(value, Enum):
                return value.value
            if isinstance(value, datetime):
                return value.isoformat()
            if hasattr(value, "item"):
                try:
                    return value.item()
                except (TypeError, ValueError):
                    pass
            return value

        def style_table(
            sheet: Any,
            header_row: int,
            *,
            freeze_cell: str | None = None,
            filter_table: bool = True,
        ) -> None:
            sheet.sheet_view.showGridLines = False
            if freeze_cell:
                sheet.freeze_panes = freeze_cell
            if filter_table and sheet.max_row >= header_row:
                sheet.auto_filter.ref = (
                    f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
                )
            for cell in sheet[header_row]:
                cell.fill = PatternFill("solid", fgColor=navy)
                cell.font = Font(color=white, bold=True)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = header_border
            sheet.row_dimensions[header_row].height = 28
            for row in sheet.iter_rows(min_row=header_row + 1):
                for cell in row:
                    cell.font = Font(color=dark_text, size=10)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for column_index in range(1, sheet.max_column + 1):
                values = [
                    str(sheet.cell(row=row, column=column_index).value or "")
                    for row in range(1, min(sheet.max_row, 250) + 1)
                ]
                width = max((len(value) for value in values), default=12) + 2
                sheet.column_dimensions[get_column_letter(column_index)].width = min(
                    max(width, 12), 52
                )

        def write_table(
            sheet: Any,
            headers: list[str],
            rows: list[list[Any]],
            *,
            header_row: int = 1,
        ) -> None:
            for column, header in enumerate(headers, 1):
                sheet.cell(header_row, column, header)
            for row_index, values in enumerate(rows, header_row + 1):
                for column_index, value in enumerate(values, 1):
                    sheet.cell(row_index, column_index, excel_value(value))
            for column_index, header in enumerate(headers, 1):
                header_lower = header.lower()
                number_format = None
                if "seconds" in header_lower or header_lower in {
                    "value",
                    "min",
                    "mean",
                    "p50",
                    "p95",
                    "p99",
                    "max",
                }:
                    number_format = "0.000"
                elif any(word in header_lower for word in ("count", "tokens")):
                    number_format = "#,##0"
                if number_format:
                    for row_index in range(header_row + 1, sheet.max_row + 1):
                        cell = sheet.cell(row_index, column_index)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = number_format

        summary_sheet = workbook.active
        summary_sheet.title = "Benchmark Summary"
        summary_sheet.merge_cells("A1:C1")
        summary_sheet["A1"] = "LLM Benchmark Summary"
        summary_sheet["A1"].fill = PatternFill("solid", fgColor=navy)
        summary_sheet["A1"].font = Font(color=white, bold=True, size=16)
        summary_sheet["A1"].alignment = Alignment(vertical="center")
        summary_sheet.row_dimensions[1].height = 30
        summary_sheet.merge_cells("A2:C2")
        summary_sheet["A2"] = (
            "Client timings, request data, and server telemetry are kept separate "
            "to preserve their measurement meaning."
        )
        summary_sheet["A2"].fill = PatternFill("solid", fgColor=pale_blue)
        summary_sheet["A2"].alignment = Alignment(wrap_text=True)
        summary_units = self.get_summary_units()
        summary_rows = [
            [key, value, summary_units.get(key, "")]
            for key, value in self.get_summary().items()
        ]
        write_table(
            summary_sheet,
            ["Metric", "Value", "Unit"],
            summary_rows,
            header_row=4,
        )
        style_table(summary_sheet, 4, freeze_cell="A5", filter_table=False)
        summary_sheet.column_dimensions["A"].width = 42
        summary_sheet.column_dimensions["B"].width = 52
        summary_sheet.column_dimensions["C"].width = 22
        for row_index in range(5, summary_sheet.max_row + 1):
            metric = str(summary_sheet.cell(row_index, 1).value or "").lower()
            value_cell = summary_sheet.cell(row_index, 2)
            if isinstance(value_cell.value, (int, float)):
                if "percent" in metric:
                    value_cell.number_format = "0.00"
                elif any(word in metric for word in ("time", "latency", "duration", "ttf", "tpot")):
                    value_cell.number_format = "0.000"
                elif metric == "concurrency" or any(
                    word in metric for word in ("request", "token")
                ):
                    value_cell.number_format = "#,##0"

        request_dicts = [metric.to_dict() for metric in self.request_metrics]
        request_headers = list(request_dicts[0]) if request_dicts else ["request_id"]
        request_rows = [[row.get(header) for header in request_headers] for row in request_dicts]
        requests_sheet = workbook.create_sheet("Requests")
        write_table(requests_sheet, request_headers, request_rows)
        style_table(requests_sheet, 1, freeze_cell="A2")

        server_summary = ((self.server_metrics or {}).get("summary") or {})
        metric_descriptions = {
            "prompt_tokens": "Prompt tokens processed during the scenario.",
            "generation_tokens": "Output tokens generated during the scenario.",
            "successful_requests": "Requests reported successful by the server.",
            "running_requests": "Concurrent requests actively executing on the server.",
            "waiting_requests": "Requests queued and waiting for execution.",
            "kv_cache_usage": "Observed fraction of KV cache in use.",
            "ttft": "Server histogram for time to first generated token.",
            "ttfb": "Server histogram for time to first response byte.",
            "tpot": "Server histogram for time per output token.",
            "e2e_latency": "Server end-to-end request latency histogram.",
            "queue_time": "Time requests spent waiting in the server queue.",
            "prefill_time": "Time spent processing prompt tokens.",
            "decode_time": "Time spent generating output tokens.",
            "inference_time": "Total server-side inference time.",
        }
        summary_rows_xlsx: list[list[Any]] = []
        for metric_name, values in (server_summary.get("counters") or {}).items():
            unit = "tokens" if "tokens" in metric_name else "count"
            summary_rows_xlsx.append([
                "counter",
                metric_name,
                "delta",
                values.get("delta"),
                unit,
                "scenario aggregate",
                values.get("source_metrics", []),
                metric_descriptions.get(metric_name, "Cumulative counter change during the scenario."),
            ])
        for metric_name, values in (server_summary.get("gauges") or {}).items():
            unit = "ratio" if metric_name == "kv_cache_usage" else "requests"
            for statistic_name in ("min", "mean", "p95", "max"):
                summary_rows_xlsx.append([
                    "gauge",
                    metric_name,
                    statistic_name,
                    values.get(statistic_name),
                    unit,
                    "scenario aggregate",
                    values.get("source_metrics", []),
                    metric_descriptions.get(metric_name, "Gauge sampled throughout the scenario."),
                ])
        for metric_name, values in (server_summary.get("histograms") or {}).items():
            for statistic_name in ("count_delta", "mean", "p50", "p95", "p99"):
                summary_rows_xlsx.append([
                    "histogram",
                    metric_name,
                    statistic_name,
                    values.get(statistic_name),
                    "requests" if statistic_name == "count_delta" else values.get("unit"),
                    "scenario aggregate",
                    values.get("source_metrics", []),
                    metric_descriptions.get(metric_name, "Server histogram observed over the scenario."),
                ])
        vllm_summary_sheet = workbook.create_sheet("vLLM Summary")
        write_table(
            vllm_summary_sheet,
            ["Category", "Metric", "Statistic", "Value", "Unit", "Scope", "Source Metrics", "Description"],
            summary_rows_xlsx,
        )
        style_table(vllm_summary_sheet, 1, freeze_cell="A2")
        for row_index in range(2, vllm_summary_sheet.max_row + 1):
            statistic = vllm_summary_sheet.cell(row_index, 3).value
            unit = vllm_summary_sheet.cell(row_index, 5).value
            value_cell = vllm_summary_sheet.cell(row_index, 4)
            if unit == "ratio":
                value_cell.number_format = "0.0%"
            elif statistic in {"delta", "count_delta"}:
                value_cell.number_format = "#,##0"

        timeline_rows: list[list[Any]] = []
        for snapshot in (self.server_metrics or {}).get("snapshots", []):
            for sample in snapshot.get("samples", []):
                timeline_rows.append([
                    snapshot.get("phase"),
                    snapshot.get("observed_at"),
                    snapshot.get("elapsed_seconds"),
                    sample.get("name"),
                    sample.get("metric_type"),
                    sample.get("labels", {}),
                    sample.get("value"),
                ])
        timeline_sheet = workbook.create_sheet("vLLM Timeline")
        write_table(
            timeline_sheet,
            ["Phase", "Observed At", "Elapsed Seconds", "Metric Name", "Metric Type", "Labels", "Value"],
            timeline_rows,
        )
        style_table(timeline_sheet, 1, freeze_cell="A2")

        during_snapshots = [
            snapshot
            for snapshot in (self.server_metrics or {}).get("snapshots", [])
            if snapshot.get("phase") == "during" and snapshot.get("elapsed_seconds") is not None
        ]
        scrape_interval = float((self.server_metrics or {}).get("scrape_interval_seconds") or 0.0)

        def contextual_gauge(snapshot: dict[str, Any], aliases: set[str], mode: str) -> float | None:
            values = [
                float(sample["value"])
                for sample in snapshot.get("samples", [])
                if sample.get("name") in aliases and sample.get("value") is not None
            ]
            if not values:
                return None
            return sum(values) if mode == "sum" else max(values)

        context_rows: list[list[Any]] = []
        context_note = (
            "Server gauges are contextual observations, not metrics attributed to this request."
        )
        for metric in self.request_metrics:
            start_elapsed = None
            end_elapsed = None
            if self.scenario_start_monotonic is not None:
                start_elapsed = metric.start_time - self.scenario_start_monotonic
                if metric.end_time is not None:
                    end_elapsed = metric.end_time - self.scenario_start_monotonic
            selected: list[dict[str, Any]] = []
            context_method = "unavailable"
            if start_elapsed is not None and end_elapsed is not None:
                selected = [
                    snapshot
                    for snapshot in during_snapshots
                    if start_elapsed <= float(snapshot["elapsed_seconds"]) <= end_elapsed
                ]
                if selected:
                    context_method = "within_request_window"
                elif during_snapshots:
                    midpoint = (start_elapsed + end_elapsed) / 2
                    nearest = min(
                        during_snapshots,
                        key=lambda snapshot: abs(float(snapshot["elapsed_seconds"]) - midpoint),
                    )
                    if abs(float(nearest["elapsed_seconds"]) - midpoint) <= max(scrape_interval, 0.05):
                        selected = [nearest]
                        context_method = "nearest_scrape"

            running = [
                value
                for snapshot in selected
                if (value := contextual_gauge(snapshot, {"vllm:num_requests_running"}, "sum")) is not None
            ]
            waiting = [
                value
                for snapshot in selected
                if (value := contextual_gauge(snapshot, {"vllm:num_requests_waiting"}, "sum")) is not None
            ]
            kv_cache = [
                value
                for snapshot in selected
                if (
                    value := contextual_gauge(
                        snapshot,
                        {"vllm:kv_cache_usage_perc", "vllm:gpu_cache_usage_perc"},
                        "max",
                    )
                ) is not None
            ]
            primary_value = metric.ttft if metric.responsiveness_metric_type == "ttft" else metric.ttfb
            context_rows.append([
                metric.request_id,
                metric.upstream_request_id,
                metric.status_code,
                metric.error.value,
                metric.response_mode,
                metric.responsiveness_metric_type,
                primary_value,
                metric.ttfb,
                metric.ttft,
                metric.total_latency,
                metric.input_tokens,
                metric.tokens_generated,
                start_elapsed,
                end_elapsed,
                context_method,
                len(selected),
                statistics.mean(running) if running else None,
                max(running) if running else None,
                statistics.mean(waiting) if waiting else None,
                max(waiting) if waiting else None,
                statistics.mean(kv_cache) if kv_cache else None,
                max(kv_cache) if kv_cache else None,
                context_note,
            ])
        context_sheet = workbook.create_sheet("Request Context")
        write_table(
            context_sheet,
            [
                "Request ID",
                "Upstream Request ID",
                "Status Code",
                "Error",
                "Response Mode",
                "Primary Responsiveness Metric",
                "Primary Responsiveness Seconds",
                "Client TTFB Seconds",
                "Client TTFT Seconds",
                "Client Latency Seconds",
                "Input Tokens",
                "Output Tokens",
                "Request Start Elapsed Seconds",
                "Request End Elapsed Seconds",
                "Context Method",
                "Server Scrape Count",
                "Running Requests Mean",
                "Running Requests Max",
                "Waiting Requests Mean",
                "Waiting Requests Max",
                "KV Cache Usage Mean",
                "KV Cache Usage Max",
                "Interpretation Note",
            ],
            context_rows,
        )
        style_table(context_sheet, 1, freeze_cell="A2")
        context_sheet.column_dimensions["W"].width = 52
        for row_index in range(2, context_sheet.max_row + 1):
            context_sheet.cell(row_index, 16).number_format = "#,##0"
            for column_index in range(17, 21):
                context_sheet.cell(row_index, column_index).number_format = "0.000"
            for column_index in range(21, 23):
                context_sheet.cell(row_index, column_index).number_format = "0.0%"

        if self.provider_monitoring:
            provider_rows: list[list[Any]] = []
            for provider_metric in self.provider_monitoring.get("metrics", []):
                for datapoint in provider_metric.get("datapoints", []):
                    values = {
                        key: value for key, value in datapoint.items() if key != "timestamp"
                    }
                    provider_rows.append([
                        provider_metric.get("metric_name"),
                        provider_metric.get("source_unit"),
                        provider_metric.get("unit"),
                        datapoint.get("timestamp"),
                        provider_metric.get("dimensions", []),
                        values,
                    ])
            provider_sheet = workbook.create_sheet("ModelArts Metrics")
            write_table(
                provider_sheet,
                ["Metric Name", "Source Unit", "Normalized Unit", "Timestamp", "Dimensions", "Values"],
                provider_rows,
            )
            style_table(provider_sheet, 1, freeze_cell="A2")

        for sheet in workbook.worksheets:
            for row_index in range(2, sheet.max_row + 1):
                if row_index % 2 == 0:
                    for cell in sheet[row_index]:
                        if cell.fill.fill_type is None:
                            cell.fill = PatternFill("solid", fgColor=pale_gray)
        workbook.save(filepath)
    
    def export_csv(self, filepath: Path) -> None:
        """Export raw metrics to CSV file."""
        if not self.request_metrics:
            return

        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        fieldnames = list(self.request_metrics[0].to_dict().keys())
        
        with open(filepath, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for metric in self.request_metrics:
                writer.writerow(metric.to_dict())

        if self.server_metrics:
            self._export_server_metrics_sidecars(filepath)
        if self.provider_monitoring:
            self._export_provider_monitoring_sidecars(filepath)

    def _export_server_metrics_sidecars(self, request_csv_path: Path) -> None:
        """Export scenario server-metrics summary and time series beside a request CSV."""
        assert self.server_metrics is not None
        json_path = request_csv_path.with_name(
            f"{request_csv_path.stem}.vllm_metrics.json"
        )
        with open(json_path, "w") as f:
            json.dump(self.server_metrics, f, indent=2)

        csv_path = request_csv_path.with_name(
            f"{request_csv_path.stem}.vllm_metrics.csv"
        )
        with open(csv_path, "w", newline="") as f:
            fieldnames = [
                "phase",
                "observed_at",
                "elapsed_seconds",
                "metric_name",
                "metric_type",
                "labels_json",
                "value",
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for snapshot in self.server_metrics.get("snapshots", []):
                for sample in snapshot.get("samples", []):
                    writer.writerow({
                        "phase": snapshot.get("phase"),
                        "observed_at": snapshot.get("observed_at"),
                        "elapsed_seconds": snapshot.get("elapsed_seconds"),
                        "metric_name": sample.get("name"),
                        "metric_type": sample.get("metric_type"),
                        "labels_json": json.dumps(sample.get("labels", {}), sort_keys=True),
                        "value": sample.get("value"),
                    })

    def _export_provider_monitoring_sidecars(self, request_csv_path: Path) -> None:
        """Export provider monitoring without mixing it into request-level CSV rows."""
        assert self.provider_monitoring is not None
        json_path = request_csv_path.with_name(
            f"{request_csv_path.stem}.modelarts_metrics.json"
        )
        with open(json_path, "w") as f:
            json.dump(self.provider_monitoring, f, indent=2)

        csv_path = request_csv_path.with_name(
            f"{request_csv_path.stem}.modelarts_metrics.csv"
        )
        with open(csv_path, "w", newline="") as f:
            fieldnames = [
                "metric_name",
                "source_unit",
                "normalized_unit",
                "timestamp",
                "dimensions_json",
                "values_json",
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for metric in self.provider_monitoring.get("metrics", []):
                for datapoint in metric.get("datapoints", []):
                    values = {
                        key: value
                        for key, value in datapoint.items()
                        if key != "timestamp"
                    }
                    writer.writerow({
                        "metric_name": metric.get("metric_name"),
                        "source_unit": metric.get("source_unit"),
                        "normalized_unit": metric.get("unit"),
                        "timestamp": datapoint.get("timestamp"),
                        "dimensions_json": json.dumps(
                            metric.get("dimensions", []), sort_keys=True
                        ),
                        "values_json": json.dumps(values, sort_keys=True),
                    })
